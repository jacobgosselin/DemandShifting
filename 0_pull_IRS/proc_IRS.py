# processing IRS data
import os
import xlrd
import pandas as pd

# Directory with downloaded files
data_dir = "/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/data/raw/irs_data_downloaded"

# ==============================================================================
# Process Table 22 files (1994-2013)
# ==============================================================================

years_table22 = range(1994, 2014)

results = []

for year in years_table22:
    file_path = os.path.join(data_dir, f"{year}_Table22.xls")

    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue

    # Read using xlrd
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)

    # Search for "Returns with net income" in column A
    net_income_row = None
    for row_idx in range(sheet.nrows):
        cell_value = str(sheet.cell_value(row_idx, 0))
        if "returns with net income" in cell_value.lower():
            net_income_row = row_idx
            break

    if net_income_row is None:
        # try to search column B if not found in column A
        # (this bites for 2004-2006, where the table format changed and the key row is in column B instead of A)
        found_in_col_b = False
        for row_idx in range(sheet.nrows):
            cell_value = str(sheet.cell_value(row_idx, 1))
            if "returns with net income" in cell_value.lower():
                net_income_row = row_idx
                # if found in column B, we need to extract total_returns and returns_with_net_income from column C
                total_returns = float(str(sheet.cell_value(row_idx-1, 2)).replace(',', ''))
                returns_with_net_income = float(str(sheet.cell_value(row_idx, 2)).replace(',', ''))
                print(f"{year}: total_returns = {total_returns:,.0f}, returns_with_net_income = {returns_with_net_income:,.0f}")
                results.append({
                    'year': year,
                    'total_returns': total_returns,
                    'returns_with_net_income': returns_with_net_income,
                    'table_number': '22',
                    'coverage': 'Returns of active corporations, other than Forms 1120-REIT, 1120-RIC, and 1120S',
                    'url': 'https://www.irs.gov/statistics/soi-tax-stats-table-21-returns-of-active-corporations-other-than-forms-1120-reit-1120-ric-and-1120s'
                })
                print("Made it here")
                found_in_col_b = True
                break
        if found_in_col_b:
            continue # skip to next year since we already found the data

    # The "Total" row should be directly above
    total_row = net_income_row - 1

    # Extract Column B (index 1) values and convert to float (removing commas)
    total_returns = float(str(sheet.cell_value(total_row, 1)).replace(',', ''))
    returns_with_net_income = float(str(sheet.cell_value(net_income_row, 1)).replace(',', ''))

    print(f"{year}: total_returns = {total_returns:,.0f}, returns_with_net_income = {returns_with_net_income:,.0f}")

    results.append({
        'year': year,
        'total_returns': total_returns,
        'returns_with_net_income': returns_with_net_income,
        'table_number': '22',
        'coverage': 'Returns of active corporations, other than Forms 1120-REIT, 1120-RIC, and 1120S',
        'url': 'https://www.irs.gov/statistics/soi-tax-stats-table-21-returns-of-active-corporations-other-than-forms-1120-reit-1120-ric-and-1120s'
    })

# ==============================================================================
# Process Table 5.3 and 5.4 files (2014-2022)
# ==============================================================================

import openpyxl

years_table53_54 = range(2014, 2023)

for year in years_table53_54:
    file_path_53 = os.path.join(data_dir, f"{year}_Table53.xlsx")
    file_path_54 = os.path.join(data_dir, f"{year}_Table54.xlsx")

    if not os.path.exists(file_path_53):
        print(f"File not found: {file_path_53}")
        continue
    if not os.path.exists(file_path_54):
        print(f"File not found: {file_path_54}")
        continue

    # Read Table 5.3 for total_returns
    wb53 = openpyxl.load_workbook(file_path_53)
    sheet53 = wb53.active

    total_returns = None
    for row_idx, row in enumerate(sheet53.iter_rows(values_only=True), start=1):
        if row[0] and "number of returns" in str(row[0]).lower():
            total_returns = float(str(row[1]).replace(',', ''))
            break

    if total_returns is None:
        print(f"Could not find 'Number of returns' in Table 5.3 for {year}")
        continue

    # Read Table 5.4 for returns_with_net_income
    wb54 = openpyxl.load_workbook(file_path_54)
    sheet54 = wb54.active

    returns_with_net_income = None
    for row_idx, row in enumerate(sheet54.iter_rows(values_only=True), start=1):
        if row[0] and "number of returns" in str(row[0]).lower():
            returns_with_net_income = float(str(row[1]).replace(',', ''))
            break

    if returns_with_net_income is None:
        print(f"Could not find 'Number of returns' in Table 5.4 for {year}")
        continue

    print(f"{year}: total_returns = {total_returns:,.0f}, returns_with_net_income = {returns_with_net_income:,.0f}")

    results.append({
        'year': year,
        'total_returns': total_returns,
        'returns_with_net_income': returns_with_net_income,
        'table_number': '5.3, 5.4',
        'coverage': 'Returns of active corporations, other than Forms 1120-REIT, 1120-RIC, and 1120S',
        'url': 'https://www.irs.gov/statistics/soi-tax-stats-table-21-returns-of-active-corporations-other-than-forms-1120-reit-1120-ric-and-1120s'
    })

# ==============================================================================
# Create final DataFrame and export to CSV
# ==============================================================================

df = pd.DataFrame(results)
print(df)
df = df.sort_values('year').reset_index(drop=True)
df['perc_neg_earnings'] = 1 - (df['returns_with_net_income'] / df['total_returns'])

print("\n=== Final Results (1994-2022) ===")
print(df)

# Export to CSV
output_path = os.path.join("/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/data/clean/", "irs_corp_returns_post94.csv")
df.to_csv(output_path, index=False)
print(f"\nSaved to: {output_path}")

# Combine with pre94 data (already constructed in clean data)
pre94_path = os.path.join("/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/data/clean/", "irs_corp_returns_pre94.csv")
pre94_df = pd.read_csv(pre94_path)
# drop table_name
pre94_df = pre94_df.drop(columns=['table_name'])
# combine
combined_df = pd.concat([pre94_df, df], ignore_index=True)
combined_df.to_csv(os.path.join("/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/data/clean/", "irs_corp_returns_combined.csv"), index=False)

# make latex table; make headers "Total Returns", "Returns with Net Income", "Percentage with No Income", "Table Number", "Coverage", "URL"
# multiply perc_neg_earnings by 100 and round to 2 decimal places
combined_df['perc_neg_earnings'] = (combined_df['perc_neg_earnings'] * 100).round(2)

# drop trailing 0s in total returns, returns with net income, and percentage with no income
combined_df['total_returns'] = combined_df['total_returns'].apply(lambda x: f"{x:,.0f}")
combined_df['returns_with_net_income'] = combined_df['returns_with_net_income'].apply(lambda x: f"{x:,.0f}")
combined_df['perc_neg_earnings'] = combined_df['perc_neg_earnings'].apply(lambda x: f"{x:.2f}".rstrip('0').rstrip('.'))
# add % sign to percentage with no income
combined_df['perc_neg_earnings'] = combined_df['perc_neg_earnings'].apply(lambda x: f"{x}\\%" if x != "0" else "0%")

combined_df = combined_df.rename(columns={
    'total_returns': 'Total Returns',
    'returns_with_net_income': 'Returns with Net Income',
    'perc_neg_earnings': '\\% without Net Income',
    'table_number': 'Table Number',
    'coverage': 'Coverage',
    'url': 'URL'
})
# drop url
combined_df = combined_df.drop(columns=['URL'])
latex_table = combined_df.to_latex(index=False, escape=False)
latex_output_path = os.path.join("/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/tables/", "irs_corp_returns_combined.tex")
# save latex table to file
with open(latex_output_path, 'w') as f:
    f.write(latex_table)
print(f"\nSaved LaTeX table to: {latex_output_path}")
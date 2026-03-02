# processing IRS data

import os
import xlrd
import pandas as pd

# Directory with downloaded files
data_dir = "/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/code/IRS_Data2/irs_data_downloaded"

# ==============================================================================
# Process Table 22 files (1994-2013)
# ==============================================================================

years_table22 = range(1994, 2014)

results = []

for year in years_table22:
    file_path = os.path.join(data_dir, f"{year}_Table22.xls")

    if not os.path.exists(file_path):
        print(f"File not found: {file_path}")
        continue

    # Read using xlrd
    workbook = xlrd.open_workbook(file_path)
    sheet = workbook.sheet_by_index(0)

    # Search for "Returns with net income" in column A
    net_income_row = None
    for row_idx in range(sheet.nrows):
        cell_value = str(sheet.cell_value(row_idx, 0))
        if "returns with net income" in cell_value.lower():
            net_income_row = row_idx
            break

    if net_income_row is None:
        print(f"Could not find 'Returns with net income' row in {year}")
        continue

    # The "Total" row should be directly above
    total_row = net_income_row - 1

    # Extract Column B (index 1) values and convert to float (removing commas)
    total_returns = float(str(sheet.cell_value(total_row, 1)).replace(',', ''))
    returns_with_net_income = float(str(sheet.cell_value(net_income_row, 1)).replace(',', ''))

    print(f"{year}: total_returns = {total_returns:,.0f}, returns_with_net_income = {returns_with_net_income:,.0f}")

    results.append({
        'year': year,
        'total_returns': total_returns,
        'returns_with_net_income': returns_with_net_income
    })

# ==============================================================================
# Process Table 5.3 and 5.4 files (2014-2022)
# ==============================================================================

import openpyxl

years_table53_54 = range(2014, 2023)

for year in years_table53_54:
    file_path_53 = os.path.join(data_dir, f"{year}_Table53.xlsx")
    file_path_54 = os.path.join(data_dir, f"{year}_Table54.xlsx")

    if not os.path.exists(file_path_53):
        print(f"File not found: {file_path_53}")
        continue
    if not os.path.exists(file_path_54):
        print(f"File not found: {file_path_54}")
        continue

    # Read Table 5.3 for total_returns
    wb53 = openpyxl.load_workbook(file_path_53)
    sheet53 = wb53.active

    total_returns = None
    for row_idx, row in enumerate(sheet53.iter_rows(values_only=True), start=1):
        if row[0] and "number of returns" in str(row[0]).lower():
            total_returns = float(str(row[1]).replace(',', ''))
            break

    if total_returns is None:
        print(f"Could not find 'Number of returns' in Table 5.3 for {year}")
        continue

    # Read Table 5.4 for returns_with_net_income
    wb54 = openpyxl.load_workbook(file_path_54)
    sheet54 = wb54.active

    returns_with_net_income = None
    for row_idx, row in enumerate(sheet54.iter_rows(values_only=True), start=1):
        if row[0] and "number of returns" in str(row[0]).lower():
            returns_with_net_income = float(str(row[1]).replace(',', ''))
            break

    if returns_with_net_income is None:
        print(f"Could not find 'Number of returns' in Table 5.4 for {year}")
        continue

    print(f"{year}: total_returns = {total_returns:,.0f}, returns_with_net_income = {returns_with_net_income:,.0f}")

    results.append({
        'year': year,
        'total_returns': total_returns,
        'returns_with_net_income': returns_with_net_income
    })

# ==============================================================================
# Create final DataFrame and export to CSV
# ==============================================================================

df = pd.DataFrame(results)
df = df.sort_values('year').reset_index(drop=True)
df['perc_neg_earnings'] = 1 - (df['returns_with_net_income'] / df['total_returns'])

print("\n=== Final Results (1994-2022) ===")
print(df)

# Export to CSV
output_path = os.path.join("/Users/jacobgosselin/Library/CloudStorage/GoogleDrive-jacob.gosselin@u.northwestern.edu/My Drive/research_ideas/negative_earnings/data/clean/", "irs_corp_returns.csv")
df.to_csv(output_path, index=False)
print(f"\nSaved to: {output_path}")